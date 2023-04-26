from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from time import sleep
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

driver = webdriver.Chrome('DRIVER AQUI')
driver.get('SITE AQUI')
sleep(5)

prices = driver.find_elements_by_class_name("PREÇOS AQUI")
names = driver.find_elements_by_class_name("PRODUTOS AQUI")

produtos = []
for nome, preco in zip(names, prices):
    nome_produto = nome.text
    preco_produto = preco.text
    produtos.append((nome_produto, preco_produto))

# Criar uma nova planilha do Excel
wb = Workbook()
ws = wb.active

# Definir a largura das colunas
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 15

# Definir estilo para o título das colunas
ws['A1'].font = Font(bold=True)
ws['B1'].font = Font(bold=True)
ws['A1'].fill = PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid')
ws['B1'].fill = PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid')
ws['A1'].alignment = Alignment(horizontal='center')
ws['B1'].alignment = Alignment(horizontal='center')

# Adicionar os dados na planilha
for nome_produto, preco_produto in produtos:
    ws.append([nome_produto, preco_produto])

# Renomear as colunas
ws.cell(row=1, column=1, value="Produtos")
ws.cell(row=1, column=2, value="Preços")

# Salvar a planilha
wb.save('produtos.xlsx')

# Fechar o navegador
driver.quit()
